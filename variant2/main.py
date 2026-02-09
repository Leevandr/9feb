import sys
import os
from collections import Counter
from datetime import datetime
import openpyxl
from PyQt6.QtWidgets import *
from PyQt6.QtGui import QPixmap
from gen.main_ui import Ui_MainForm
from gen.auth_ui import Ui_AuthForm
from gen.buyer_ui import Ui_BuyerForm
from gen.manager_role_ui import Ui_ManagerRoleForm
from gen.admin_ui import Ui_AdminForm

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Общие данные (in-memory, разделяемые между окнами)
orders = []       # [{id, items: [{name, price}], total, status, date}]
users_list = [
    {'login': 'buyer', 'password': '1', 'role': 'buyer'},
    {'login': 'manager', 'password': '1', 'role': 'manager'},
    {'login': 'admin', 'password': '1', 'role': 'admin'},
]


# ============================================================
#  Вспомогательные функции для работы с xlsx
# ============================================================
def load_products():
    path = os.path.join(BASE_DIR, 'products.xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        products.append({
            'name': str(row[0]),
            'article': str(row[1]),
            'category': str(row[2]),
            'price': float(row[3]),
            'desc': str(row[4]),
            'photo': str(row[5]) if row[5] else ''
        })
    return products


def save_products(products):
    path = os.path.join(BASE_DIR, 'products.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Наименование', 'Артикул', 'Категория', 'Цена', 'Описание', 'Фото'])
    for p in products:
        ws.append([p['name'], p['article'], p['category'],
                   p['price'], p['desc'], p['photo']])
    wb.save(path)


# ============================================================
#  ГЛАВНОЕ ОКНО  (неавторизованный пользователь)
# ============================================================
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainForm()
        self.ui.setupUi(self)

        self.products = load_products()
        self.fill_categories()
        self.fill_table(self.products)

        self.ui.search_btn.clicked.connect(self.search)
        self.ui.reset_btn.clicked.connect(self.reset)
        self.ui.auth_btn.clicked.connect(self.open_auth)
        self.ui.table.doubleClicked.connect(self.show_photo)
        self.ui.table.setSortingEnabled(True)

    def fill_categories(self):
        cats = sorted(set(p['category'] for p in self.products))
        self.ui.category_combo.clear()
        self.ui.category_combo.addItem('Все')
        self.ui.category_combo.addItems(cats)

    def fill_table(self, data):
        self.ui.table.setSortingEnabled(False)
        self.ui.table.setRowCount(len(data))
        self.ui.table.setColumnCount(5)
        self.ui.table.setHorizontalHeaderLabels(
            ['Наименование', 'Артикул', 'Категория', 'Цена', 'Фото'])
        for i, p in enumerate(data):
            self.ui.table.setItem(i, 0, QTableWidgetItem(p['name']))
            self.ui.table.setItem(i, 1, QTableWidgetItem(p['article']))
            self.ui.table.setItem(i, 2, QTableWidgetItem(p['category']))
            self.ui.table.setItem(i, 3, QTableWidgetItem(str(p['price'])))
            self.ui.table.setItem(i, 4, QTableWidgetItem(p['photo']))
        self.ui.table.setSortingEnabled(True)

    def search(self):
        text = self.ui.search_edit.text().strip().lower()
        category = self.ui.category_combo.currentText()
        p_min = self.ui.price_min.value()
        p_max = self.ui.price_max.value()
        result = self.products
        if text:
            result = [p for p in result
                      if text in p['name'].lower() or text in p['article'].lower()]
        if category != 'Все':
            result = [p for p in result if p['category'] == category]
        if p_max > 0:
            result = [p for p in result if p_min <= p['price'] <= p_max]
        self.fill_table(result)

    def reset(self):
        self.ui.search_edit.clear()
        self.ui.category_combo.setCurrentIndex(0)
        self.ui.price_min.setValue(0)
        self.ui.price_max.setValue(0)
        self.fill_table(self.products)

    def show_photo(self, index):
        photo_path = self.ui.table.item(index.row(), 4)
        if photo_path and photo_path.text():
            dlg = QDialog(self)
            dlg.setWindowTitle('Фото товара')
            lbl = QLabel()
            pix = QPixmap(os.path.join(BASE_DIR, photo_path.text()))
            lbl.setPixmap(pix.scaled(400, 400))
            lay = QVBoxLayout(dlg)
            lay.addWidget(lbl)
            dlg.exec()

    def open_auth(self):
        self.auth_win = AuthWindow()
        self.auth_win.show()


# ============================================================
#  АВТОРИЗАЦИЯ
# ============================================================
class AuthWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_AuthForm()
        self.ui.setupUi(self)
        self.ui.login_btn.clicked.connect(self.do_login)

    def do_login(self):
        login = self.ui.login_edit.text().strip()
        password = self.ui.password_edit.text().strip()

        role = None
        for u in users_list:
            if u['login'] == login and u['password'] == password:
                role = u['role']
                break

        if role is None:
            QMessageBox.warning(self, 'Ошибка', 'Неверный логин или пароль')
            return

        if role == 'buyer':
            self.w = BuyerWindow(login)
        elif role == 'manager':
            self.w = ManagerRoleWindow()
        else:
            self.w = AdminWindow()
        self.w.show()
        self.close()


# ============================================================
#  ОКНО ПОКУПАТЕЛЯ  (каталог, корзина, заказы, личный кабинет)
# ============================================================
class BuyerWindow(QWidget):
    def __init__(self, username='buyer'):
        super().__init__()
        self.ui = Ui_BuyerForm()
        self.ui.setupUi(self)
        self.username = username
        self.products = load_products()
        self.cart = []  # [{name, article, price}]

        # --- Каталог ---
        self.fill_categories()
        self.fill_catalog(self.products)
        self.ui.search_btn.clicked.connect(self.search)
        self.ui.reset_btn.clicked.connect(self.reset_search)
        self.ui.catalog_table.doubleClicked.connect(self.show_photo)
        self.ui.catalog_table.setSortingEnabled(True)
        self.ui.add_to_cart_btn.clicked.connect(self.add_to_cart)

        # --- Корзина ---
        self.ui.remove_from_cart_btn.clicked.connect(self.remove_from_cart)
        self.ui.checkout_btn.clicked.connect(self.checkout)

        # --- Профиль ---
        self.ui.profile_name.setText(f'Покупатель: {self.username}')
        self.refresh_orders()

    # --- Каталог ---
    def fill_categories(self):
        cats = sorted(set(p['category'] for p in self.products))
        self.ui.category_combo.clear()
        self.ui.category_combo.addItem('Все')
        self.ui.category_combo.addItems(cats)

    def fill_catalog(self, data):
        self.ui.catalog_table.setSortingEnabled(False)
        self.ui.catalog_table.setRowCount(len(data))
        self.ui.catalog_table.setColumnCount(5)
        self.ui.catalog_table.setHorizontalHeaderLabels(
            ['Наименование', 'Артикул', 'Категория', 'Цена', 'Фото'])
        for i, p in enumerate(data):
            self.ui.catalog_table.setItem(i, 0, QTableWidgetItem(p['name']))
            self.ui.catalog_table.setItem(i, 1, QTableWidgetItem(p['article']))
            self.ui.catalog_table.setItem(i, 2, QTableWidgetItem(p['category']))
            self.ui.catalog_table.setItem(i, 3, QTableWidgetItem(str(p['price'])))
            self.ui.catalog_table.setItem(i, 4, QTableWidgetItem(p['photo']))
        self.ui.catalog_table.setSortingEnabled(True)

    def search(self):
        text = self.ui.search_edit.text().strip().lower()
        cat = self.ui.category_combo.currentText()
        result = self.products
        if text:
            result = [p for p in result
                      if text in p['name'].lower() or text in p['article'].lower()]
        if cat != 'Все':
            result = [p for p in result if p['category'] == cat]
        self.fill_catalog(result)

    def reset_search(self):
        self.ui.search_edit.clear()
        self.ui.category_combo.setCurrentIndex(0)
        self.fill_catalog(self.products)

    def show_photo(self, index):
        photo_path = self.ui.catalog_table.item(index.row(), 4)
        if photo_path and photo_path.text():
            dlg = QDialog(self)
            dlg.setWindowTitle('Фото товара')
            lbl = QLabel()
            pix = QPixmap(os.path.join(BASE_DIR, photo_path.text()))
            lbl.setPixmap(pix.scaled(400, 400))
            lay = QVBoxLayout(dlg)
            lay.addWidget(lbl)
            dlg.exec()

    def add_to_cart(self):
        row = self.ui.catalog_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, 'Внимание', 'Выберите товар')
            return
        name = self.ui.catalog_table.item(row, 0).text()
        article = self.ui.catalog_table.item(row, 1).text()
        price = float(self.ui.catalog_table.item(row, 3).text())
        self.cart.append({'name': name, 'article': article, 'price': price})
        self.refresh_cart()
        QMessageBox.information(self, 'Корзина', f'{name} добавлен в корзину')

    # --- Корзина ---
    def refresh_cart(self):
        self.ui.cart_table.setRowCount(len(self.cart))
        self.ui.cart_table.setColumnCount(3)
        self.ui.cart_table.setHorizontalHeaderLabels(
            ['Наименование', 'Артикул', 'Цена'])
        for i, item in enumerate(self.cart):
            self.ui.cart_table.setItem(i, 0, QTableWidgetItem(item['name']))
            self.ui.cart_table.setItem(i, 1, QTableWidgetItem(item['article']))
            self.ui.cart_table.setItem(i, 2, QTableWidgetItem(str(item['price'])))
        total = sum(item['price'] for item in self.cart)
        self.ui.cart_total_label.setText(f'Итого: {total}')

    def remove_from_cart(self):
        row = self.ui.cart_table.currentRow()
        if row >= 0:
            self.cart.pop(row)
            self.refresh_cart()

    def checkout(self):
        if not self.cart:
            QMessageBox.warning(self, 'Внимание', 'Корзина пуста')
            return
        total = sum(item['price'] for item in self.cart)
        order = {
            'id': len(orders) + 1,
            'buyer': self.username,
            'items': list(self.cart),
            'total': total,
            'status': 'Новый',
            'date': datetime.now().strftime('%Y-%m-%d %H:%M')
        }
        orders.append(order)
        self.cart.clear()
        self.refresh_cart()
        self.refresh_orders()
        QMessageBox.information(self, 'Заказ',
                                f'Заказ #{order["id"]} оформлен на сумму {total}')

    # --- Заказы ---
    def refresh_orders(self):
        my_orders = [o for o in orders if o['buyer'] == self.username]
        self.ui.orders_table.setRowCount(len(my_orders))
        self.ui.orders_table.setColumnCount(4)
        self.ui.orders_table.setHorizontalHeaderLabels(
            ['№', 'Дата', 'Сумма', 'Статус'])
        for i, o in enumerate(my_orders):
            self.ui.orders_table.setItem(i, 0, QTableWidgetItem(str(o['id'])))
            self.ui.orders_table.setItem(i, 1, QTableWidgetItem(o['date']))
            self.ui.orders_table.setItem(i, 2, QTableWidgetItem(str(o['total'])))
            self.ui.orders_table.setItem(i, 3, QTableWidgetItem(o['status']))
        self.ui.profile_orders_count.setText(f'Заказов: {len(my_orders)}')


# ============================================================
#  ОКНО МЕНЕДЖЕРА  (CRUD товаров, обработка заказов)
# ============================================================
class ManagerRoleWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_ManagerRoleForm()
        self.ui.setupUi(self)
        self.products = load_products()

        # --- Товары ---
        self.fill_product_table()
        self.ui.product_table.clicked.connect(self.on_product_click)
        self.ui.add_btn.clicked.connect(self.add_product)
        self.ui.edit_btn.clicked.connect(self.edit_product)
        self.ui.delete_btn.clicked.connect(self.delete_product)

        # --- Заказы ---
        self.refresh_orders()
        self.ui.accept_btn.clicked.connect(lambda: self.change_status('В обработке'))
        self.ui.complete_btn.clicked.connect(lambda: self.change_status('Выполнен'))
        self.ui.cancel_btn.clicked.connect(lambda: self.change_status('Отменён'))

    # --- Товары CRUD ---
    def fill_product_table(self):
        self.ui.product_table.setRowCount(len(self.products))
        self.ui.product_table.setColumnCount(6)
        self.ui.product_table.setHorizontalHeaderLabels(
            ['Наименование', 'Артикул', 'Категория', 'Цена', 'Описание', 'Фото'])
        for i, p in enumerate(self.products):
            self.ui.product_table.setItem(i, 0, QTableWidgetItem(p['name']))
            self.ui.product_table.setItem(i, 1, QTableWidgetItem(p['article']))
            self.ui.product_table.setItem(i, 2, QTableWidgetItem(p['category']))
            self.ui.product_table.setItem(i, 3, QTableWidgetItem(str(p['price'])))
            self.ui.product_table.setItem(i, 4, QTableWidgetItem(p['desc']))
            self.ui.product_table.setItem(i, 5, QTableWidgetItem(p['photo']))

    def on_product_click(self, index):
        row = index.row()
        self.ui.name_edit.setText(self.ui.product_table.item(row, 0).text())
        self.ui.article_edit.setText(self.ui.product_table.item(row, 1).text())
        self.ui.category_edit.setText(self.ui.product_table.item(row, 2).text())
        self.ui.price_edit.setText(self.ui.product_table.item(row, 3).text())
        self.ui.desc_edit.setText(self.ui.product_table.item(row, 4).text())
        self.ui.photo_edit.setText(self.ui.product_table.item(row, 5).text())

    def add_product(self):
        name = self.ui.name_edit.text().strip()
        article = self.ui.article_edit.text().strip()
        if not name or not article:
            QMessageBox.warning(self, 'Ошибка',
                                'Заполните наименование и артикул')
            return
        self.products.append({
            'name': name,
            'article': article,
            'category': self.ui.category_edit.text().strip(),
            'price': float(self.ui.price_edit.text() or 0),
            'desc': self.ui.desc_edit.text().strip(),
            'photo': self.ui.photo_edit.text().strip()
        })
        save_products(self.products)
        self.fill_product_table()
        self.clear_form()
        QMessageBox.information(self, 'Успех', 'Товар добавлен')

    def edit_product(self):
        row = self.ui.product_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, 'Ошибка', 'Выберите товар')
            return
        self.products[row] = {
            'name': self.ui.name_edit.text().strip(),
            'article': self.ui.article_edit.text().strip(),
            'category': self.ui.category_edit.text().strip(),
            'price': float(self.ui.price_edit.text() or 0),
            'desc': self.ui.desc_edit.text().strip(),
            'photo': self.ui.photo_edit.text().strip()
        }
        save_products(self.products)
        self.fill_product_table()
        QMessageBox.information(self, 'Успех', 'Товар обновлён')

    def delete_product(self):
        row = self.ui.product_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, 'Ошибка', 'Выберите товар')
            return
        reply = QMessageBox.question(
            self, 'Подтверждение',
            f'Удалить "{self.products[row]["name"]}"?')
        if reply == QMessageBox.StandardButton.Yes:
            self.products.pop(row)
            save_products(self.products)
            self.fill_product_table()
            self.clear_form()

    def clear_form(self):
        self.ui.name_edit.clear()
        self.ui.article_edit.clear()
        self.ui.category_edit.clear()
        self.ui.price_edit.clear()
        self.ui.desc_edit.clear()
        self.ui.photo_edit.clear()

    # --- Заказы ---
    def refresh_orders(self):
        self.ui.orders_table.setRowCount(len(orders))
        self.ui.orders_table.setColumnCount(5)
        self.ui.orders_table.setHorizontalHeaderLabels(
            ['№', 'Покупатель', 'Дата', 'Сумма', 'Статус'])
        for i, o in enumerate(orders):
            self.ui.orders_table.setItem(i, 0, QTableWidgetItem(str(o['id'])))
            self.ui.orders_table.setItem(i, 1, QTableWidgetItem(o['buyer']))
            self.ui.orders_table.setItem(i, 2, QTableWidgetItem(o['date']))
            self.ui.orders_table.setItem(i, 3, QTableWidgetItem(str(o['total'])))
            self.ui.orders_table.setItem(i, 4, QTableWidgetItem(o['status']))

    def change_status(self, new_status):
        row = self.ui.orders_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, 'Ошибка', 'Выберите заказ')
            return
        orders[row]['status'] = new_status
        self.refresh_orders()


# ============================================================
#  ОКНО АДМИНИСТРАТОРА  (пользователи, статистика продаж)
# ============================================================
class AdminWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_AdminForm()
        self.ui.setupUi(self)

        # --- Пользователи ---
        self.refresh_users()
        self.ui.add_user_btn.clicked.connect(self.add_user)
        self.ui.delete_user_btn.clicked.connect(self.delete_user)

        # --- Статистика ---
        self.refresh_stats()

    # --- Пользователи ---
    def refresh_users(self):
        self.ui.users_table.setRowCount(len(users_list))
        self.ui.users_table.setColumnCount(3)
        self.ui.users_table.setHorizontalHeaderLabels(
            ['Логин', 'Пароль', 'Роль'])
        for i, u in enumerate(users_list):
            self.ui.users_table.setItem(i, 0, QTableWidgetItem(u['login']))
            self.ui.users_table.setItem(i, 1, QTableWidgetItem(u['password']))
            self.ui.users_table.setItem(i, 2, QTableWidgetItem(u['role']))

    def add_user(self):
        login = self.ui.login_edit.text().strip()
        password = self.ui.password_input.text().strip()
        role = self.ui.role_combo.currentText()
        if not login or not password:
            QMessageBox.warning(self, 'Ошибка', 'Заполните логин и пароль')
            return
        users_list.append({'login': login, 'password': password, 'role': role})
        self.refresh_users()
        self.ui.login_edit.clear()
        self.ui.password_input.clear()
        QMessageBox.information(self, 'Успех', f'Пользователь {login} добавлен')

    def delete_user(self):
        row = self.ui.users_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, 'Ошибка', 'Выберите пользователя')
            return
        reply = QMessageBox.question(
            self, 'Подтверждение',
            f'Удалить "{users_list[row]["login"]}"?')
        if reply == QMessageBox.StandardButton.Yes:
            users_list.pop(row)
            self.refresh_users()

    # --- Статистика продаж ---
    def refresh_stats(self):
        self.ui.total_orders_label.setText(f'Всего заказов: {len(orders)}')
        total_revenue = sum(o['total'] for o in orders)
        self.ui.total_revenue_label.setText(f'Общая выручка: {total_revenue}')

        # Статистика по категориям
        cat_revenue = Counter()
        for o in orders:
            for item in o['items']:
                cat_revenue[item.get('name', 'Без имени')] += item['price']

        self.ui.stats_table.setRowCount(len(cat_revenue))
        self.ui.stats_table.setColumnCount(2)
        self.ui.stats_table.setHorizontalHeaderLabels(['Товар', 'Выручка'])
        for i, (name, rev) in enumerate(cat_revenue.most_common()):
            self.ui.stats_table.setItem(i, 0, QTableWidgetItem(name))
            self.ui.stats_table.setItem(i, 1, QTableWidgetItem(str(rev)))


# ============================================================
#  ЗАПУСК
# ============================================================
if __name__ == '__main__':
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())
