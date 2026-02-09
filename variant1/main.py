import sys
import os
from collections import Counter
import openpyxl
from PyQt6.QtWidgets import *
from PyQt6.QtGui import QPixmap
from gen.main_ui import Ui_MainForm
from gen.auth_ui import Ui_AuthForm
from gen.reader_ui import Ui_ReaderForm
from gen.employee_ui import Ui_EmployeeForm
from gen.manager_role_ui import Ui_ManagerRoleForm

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


# ============================================================
#  Вспомогательные функции для работы с xlsx
# ============================================================
def load_books():
    path = os.path.join(BASE_DIR, 'books.xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    books = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        books.append({
            'title': str(row[0]),
            'author': str(row[1]),
            'genre': str(row[2]),
            'photo': str(row[3]) if row[3] else ''
        })
    return books


def save_books(books):
    path = os.path.join(BASE_DIR, 'books.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Название', 'Автор', 'Жанр', 'Фото'])
    for b in books:
        ws.append([b['title'], b['author'], b['genre'], b['photo']])
    wb.save(path)


# ============================================================
#  ГЛАВНОЕ ОКНО  (неавторизованный пользователь)
# ============================================================
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainForm()
        self.ui.setupUi(self)

        self.books = load_books()
        self.fill_genres()
        self.fill_table(self.books)

        self.ui.search_btn.clicked.connect(self.search)
        self.ui.reset_btn.clicked.connect(self.reset)
        self.ui.auth_btn.clicked.connect(self.open_auth)
        self.ui.table.doubleClicked.connect(self.show_photo)
        self.ui.table.setSortingEnabled(True)

    def fill_genres(self):
        genres = sorted(set(b['genre'] for b in self.books))
        self.ui.genre_combo.clear()
        self.ui.genre_combo.addItem('Все')
        self.ui.genre_combo.addItems(genres)

    def fill_table(self, data):
        self.ui.table.setSortingEnabled(False)
        self.ui.table.setRowCount(len(data))
        self.ui.table.setColumnCount(4)
        self.ui.table.setHorizontalHeaderLabels(
            ['Название', 'Автор', 'Жанр', 'Фото'])
        for i, b in enumerate(data):
            self.ui.table.setItem(i, 0, QTableWidgetItem(b['title']))
            self.ui.table.setItem(i, 1, QTableWidgetItem(b['author']))
            self.ui.table.setItem(i, 2, QTableWidgetItem(b['genre']))
            self.ui.table.setItem(i, 3, QTableWidgetItem(b['photo']))
        self.ui.table.setSortingEnabled(True)

    def search(self):
        author = self.ui.author_edit.text().strip().lower()
        genre = self.ui.genre_combo.currentText()
        result = self.books
        if author:
            result = [b for b in result if author in b['author'].lower()]
        if genre != 'Все':
            result = [b for b in result if b['genre'] == genre]
        self.fill_table(result)

    def reset(self):
        self.ui.author_edit.clear()
        self.ui.genre_combo.setCurrentIndex(0)
        self.fill_table(self.books)

    def show_photo(self, index):
        photo_path = self.ui.table.item(index.row(), 3)
        if photo_path and photo_path.text():
            dlg = QDialog(self)
            dlg.setWindowTitle('Обложка')
            lbl = QLabel()
            pix = QPixmap(os.path.join(BASE_DIR, photo_path.text()))
            lbl.setPixmap(pix.scaled(400, 400))
            lay = QVBoxLayout(dlg)
            lay.addWidget(lbl)
            dlg.exec()

    def open_auth(self):
        self.auth_win = AuthWindow()
        self.auth_win.show()


# ============================================================
#  АВТОРИЗАЦИЯ
# ============================================================
class AuthWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_AuthForm()
        self.ui.setupUi(self)
        self.ui.login_btn.clicked.connect(self.do_login)

    def do_login(self):
        login = self.ui.login_edit.text().strip()
        password = self.ui.password_edit.text().strip()

        users = {
            ('reader', '1'): 'reader',
            ('employee', '1'): 'employee',
            ('manager', '1'): 'manager',
        }
        role = users.get((login, password))
        if role is None:
            QMessageBox.warning(self, 'Ошибка', 'Неверный логин или пароль')
            return

        if role == 'reader':
            self.w = ReaderWindow()
        elif role == 'employee':
            self.w = EmployeeWindow()
        else:
            self.w = ManagerRoleWindow()
        self.w.show()
        self.close()


# ============================================================
#  ОКНО ЧИТАТЕЛЯ  (поиск, чтение, подписка)
# ============================================================
class ReaderWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_ReaderForm()
        self.ui.setupUi(self)

        self.books = load_books()
        self.subscriptions = []

        # --- Поиск ---
        self.fill_genres()
        self.fill_table(self.books)
        self.ui.search_btn.clicked.connect(self.search)
        self.ui.reset_btn.clicked.connect(self.reset)
        self.ui.table.doubleClicked.connect(self.show_photo)
        self.ui.table.setSortingEnabled(True)

        # --- Чтение ---
        self.ui.read_btn.clicked.connect(self.read_book)

        # --- Подписки ---
        genres = sorted(set(b['genre'] for b in self.books))
        self.ui.sub_genre_combo.addItems(genres)
        self.ui.subscribe_btn.clicked.connect(self.subscribe)
        self.ui.unsubscribe_btn.clicked.connect(self.unsubscribe)

    # --- Поиск (аналогично главному окну) ---
    def fill_genres(self):
        genres = sorted(set(b['genre'] for b in self.books))
        self.ui.genre_combo.clear()
        self.ui.genre_combo.addItem('Все')
        self.ui.genre_combo.addItems(genres)

    def fill_table(self, data):
        self.ui.table.setSortingEnabled(False)
        self.ui.table.setRowCount(len(data))
        self.ui.table.setColumnCount(4)
        self.ui.table.setHorizontalHeaderLabels(
            ['Название', 'Автор', 'Жанр', 'Фото'])
        for i, b in enumerate(data):
            self.ui.table.setItem(i, 0, QTableWidgetItem(b['title']))
            self.ui.table.setItem(i, 1, QTableWidgetItem(b['author']))
            self.ui.table.setItem(i, 2, QTableWidgetItem(b['genre']))
            self.ui.table.setItem(i, 3, QTableWidgetItem(b['photo']))
        self.ui.table.setSortingEnabled(True)

    def search(self):
        author = self.ui.author_edit.text().strip().lower()
        genre = self.ui.genre_combo.currentText()
        result = self.books
        if author:
            result = [b for b in result if author in b['author'].lower()]
        if genre != 'Все':
            result = [b for b in result if b['genre'] == genre]
        self.fill_table(result)

    def reset(self):
        self.ui.author_edit.clear()
        self.ui.genre_combo.setCurrentIndex(0)
        self.fill_table(self.books)

    def show_photo(self, index):
        photo_path = self.ui.table.item(index.row(), 3)
        if photo_path and photo_path.text():
            dlg = QDialog(self)
            dlg.setWindowTitle('Обложка')
            lbl = QLabel()
            pix = QPixmap(os.path.join(BASE_DIR, photo_path.text()))
            lbl.setPixmap(pix.scaled(400, 400))
            lay = QVBoxLayout(dlg)
            lay.addWidget(lbl)
            dlg.exec()

    # --- Чтение ---
    def read_book(self):
        row = self.ui.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, 'Внимание', 'Выберите книгу в таблице')
            return
        title = self.ui.table.item(row, 0).text()
        author = self.ui.table.item(row, 1).text()
        genre = self.ui.table.item(row, 2).text()
        self.ui.book_title_label.setText(f'{title} — {author}')
        self.ui.text_edit.setPlainText(
            f'Название: {title}\nАвтор: {author}\nЖанр: {genre}\n\n'
            f'Здесь отображается содержимое книги "{title}"...')
        self.ui.tabs.setCurrentIndex(1)

    # --- Подписки ---
    def subscribe(self):
        genre = self.ui.sub_genre_combo.currentText()
        if genre and genre not in self.subscriptions:
            self.subscriptions.append(genre)
            self.ui.sub_list.addItem(genre)
            QMessageBox.information(self, 'Подписка',
                                    f'Вы подписались на жанр "{genre}"')

    def unsubscribe(self):
        item = self.ui.sub_list.currentItem()
        if item:
            self.subscriptions.remove(item.text())
            self.ui.sub_list.takeItem(self.ui.sub_list.row(item))


# ============================================================
#  ОКНО СОТРУДНИКА  (CRUD книг, фото)
# ============================================================
class EmployeeWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_EmployeeForm()
        self.ui.setupUi(self)

        self.books = load_books()
        self.fill_table()

        self.ui.table.clicked.connect(self.on_row_click)
        self.ui.add_btn.clicked.connect(self.add_book)
        self.ui.edit_btn.clicked.connect(self.edit_book)
        self.ui.delete_btn.clicked.connect(self.delete_book)
        self.ui.photo_browse_btn.clicked.connect(self.browse_photo)

    def fill_table(self):
        self.ui.table.setSortingEnabled(False)
        self.ui.table.setRowCount(len(self.books))
        self.ui.table.setColumnCount(4)
        self.ui.table.setHorizontalHeaderLabels(
            ['Название', 'Автор', 'Жанр', 'Фото'])
        for i, b in enumerate(self.books):
            self.ui.table.setItem(i, 0, QTableWidgetItem(b['title']))
            self.ui.table.setItem(i, 1, QTableWidgetItem(b['author']))
            self.ui.table.setItem(i, 2, QTableWidgetItem(b['genre']))
            self.ui.table.setItem(i, 3, QTableWidgetItem(b['photo']))
        self.ui.table.setSortingEnabled(True)

    def on_row_click(self, index):
        row = index.row()
        self.ui.title_edit.setText(self.ui.table.item(row, 0).text())
        self.ui.author_edit.setText(self.ui.table.item(row, 1).text())
        self.ui.genre_edit.setText(self.ui.table.item(row, 2).text())
        self.ui.photo_edit.setText(self.ui.table.item(row, 3).text())

    def add_book(self):
        title = self.ui.title_edit.text().strip()
        author = self.ui.author_edit.text().strip()
        genre = self.ui.genre_edit.text().strip()
        photo = self.ui.photo_edit.text().strip()
        if not title or not author or not genre:
            QMessageBox.warning(self, 'Ошибка', 'Заполните название, автора и жанр')
            return
        self.books.append({
            'title': title, 'author': author,
            'genre': genre, 'photo': photo
        })
        save_books(self.books)
        self.fill_table()
        self.clear_form()
        QMessageBox.information(self, 'Успех', 'Книга добавлена')

    def edit_book(self):
        row = self.ui.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, 'Ошибка', 'Выберите книгу для редактирования')
            return
        self.books[row] = {
            'title': self.ui.title_edit.text().strip(),
            'author': self.ui.author_edit.text().strip(),
            'genre': self.ui.genre_edit.text().strip(),
            'photo': self.ui.photo_edit.text().strip()
        }
        save_books(self.books)
        self.fill_table()
        QMessageBox.information(self, 'Успех', 'Книга обновлена')

    def delete_book(self):
        row = self.ui.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, 'Ошибка', 'Выберите книгу для удаления')
            return
        reply = QMessageBox.question(self, 'Подтверждение',
                                     f'Удалить "{self.books[row]["title"]}"?')
        if reply == QMessageBox.StandardButton.Yes:
            self.books.pop(row)
            save_books(self.books)
            self.fill_table()
            self.clear_form()

    def browse_photo(self):
        path, _ = QFileDialog.getOpenFileName(
            self, 'Выберите фото', BASE_DIR,
            'Изображения (*.png *.jpg *.jpeg *.bmp)')
        if path:
            self.ui.photo_edit.setText(path)

    def clear_form(self):
        self.ui.title_edit.clear()
        self.ui.author_edit.clear()
        self.ui.genre_edit.clear()
        self.ui.photo_edit.clear()


# ============================================================
#  ОКНО РУКОВОДИТЕЛЯ  (статистика, доп. материалы)
# ============================================================
class ManagerRoleWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_ManagerRoleForm()
        self.ui.setupUi(self)

        self.books = load_books()
        self.fill_stats()
        self.fill_materials()

    def fill_stats(self):
        # Общее количество
        self.ui.total_label.setText(f'Всего книг: {len(self.books)}')

        # По жанрам
        genre_counts = Counter(b['genre'] for b in self.books)
        self.ui.genre_table.setRowCount(len(genre_counts))
        self.ui.genre_table.setColumnCount(2)
        self.ui.genre_table.setHorizontalHeaderLabels(['Жанр', 'Количество'])
        for i, (genre, count) in enumerate(genre_counts.most_common()):
            self.ui.genre_table.setItem(i, 0, QTableWidgetItem(genre))
            self.ui.genre_table.setItem(i, 1, QTableWidgetItem(str(count)))

        # По авторам
        author_counts = Counter(b['author'] for b in self.books)
        self.ui.author_table.setRowCount(len(author_counts))
        self.ui.author_table.setColumnCount(2)
        self.ui.author_table.setHorizontalHeaderLabels(['Автор', 'Количество'])
        for i, (author, count) in enumerate(author_counts.most_common()):
            self.ui.author_table.setItem(i, 0, QTableWidgetItem(author))
            self.ui.author_table.setItem(i, 1, QTableWidgetItem(str(count)))

    def fill_materials(self):
        # Список фото-материалов из книг
        photos = [b for b in self.books if b['photo']]
        self.ui.mat_table.setRowCount(len(photos))
        self.ui.mat_table.setColumnCount(3)
        self.ui.mat_table.setHorizontalHeaderLabels(
            ['Книга', 'Автор', 'Файл материала'])
        for i, b in enumerate(photos):
            self.ui.mat_table.setItem(i, 0, QTableWidgetItem(b['title']))
            self.ui.mat_table.setItem(i, 1, QTableWidgetItem(b['author']))
            self.ui.mat_table.setItem(i, 2, QTableWidgetItem(b['photo']))


# ============================================================
#  ЗАПУСК
# ============================================================
if __name__ == '__main__':
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())
